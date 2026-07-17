"""Genera la plantilla de Excel para importar en Routine App.

Uso:  python tools/generate_template.py
Produce: docs/Plantilla_Rutina.xlsx

Hojas: Horario, Tareas, Ejercicios, Hitos. Los encabezados coinciden con lo que
reconoce el importador (TemplateImporter.kt), que identifica cada campo por
palabras clave, así que puedes reordenar o renombrar cerca del original.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
import os

HEADER_FILL = PatternFill("solid", fgColor="4A7A5A")
HEADER_FONT = Font(bold=True, color="FFFFFF")
THIN = Side(style="thin", color="D0D7DE")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DIAS = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado", "Domingo"]
CONTEXTOS_H = ["Estudio", "Calistenia", "General"]
CONTEXTOS = ["Estudio", "Calistenia"]
ESTADOS = ["Hecho", "Actual", "Pendiente"]


def style_headers(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w


def add_list(ws, options, col_letter, last_row=100):
    dv = DataValidation(type="list", formula1='"%s"' % ",".join(options), allow_blank=True)
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}2:{col_letter}{last_row}")


def text_col(ws, col, last=60):
    for row in range(2, last):
        ws.cell(row=row, column=col).number_format = "@"


wb = Workbook()

# --- Hoja 1: Horario (timeline del día) ---
ws = wb.active
ws.title = "Horario"
ws.append(["Día", "Inicio", "Fin", "Actividad", "Contexto", "Notas"])
for r in [
    ["Jueves", "05:30", "", "Despertar", "General", ""],
    ["Jueves", "06:00", "", "Desayuno", "General", ""],
    ["Jueves", "07:00", "14:00", "Clases", "Estudio", "Foco de la mañana"],
    ["Jueves", "14:30", "", "Comida — fuerte", "General", ""],
    ["Jueves", "16:30", "17:30", "Entrenamiento", "Calistenia", ""],
    ["Jueves", "21:30", "", "Dormir", "General", ""],
    ["Lunes", "07:00", "14:00", "Clases", "Estudio", ""],
    ["Lunes", "18:00", "19:00", "Entrenamiento", "Calistenia", ""],
]:
    ws.append(r)
style_headers(ws, 6)
autosize(ws, [12, 9, 9, 26, 13, 22])
add_list(ws, DIAS, "A")
add_list(ws, CONTEXTOS_H, "E")

# --- Hoja 2: Tareas (pestaña Hoy de Estudio) ---
ws = wb.create_sheet("Tareas")
ws.append(["Contexto", "Día", "Tarea", "Estimación", "Orden"])
for r in [
    ["Estudio", "Jueves", "Wireframes pantallas", "2h", 1],
    ["Estudio", "Jueves", "Revisión con mentor", "45m", 2],
    ["Estudio", "Jueves", "Ajustar propuesta", "1h", 3],
    ["Estudio", "Lunes", "Leer capítulo 3", "1h", 1],
]:
    ws.append(r)
style_headers(ws, 5)
autosize(ws, [13, 12, 28, 12, 8])
add_list(ws, CONTEXTOS, "A")
add_list(ws, DIAS, "B")

# --- Hoja 3: Ejercicios (pestaña Hoy de Calistenia) ---
ws = wb.create_sheet("Ejercicios")
ws.append(["Día", "Hora", "Ejercicio", "Series", "Reps", "Notas"])
for r in [
    ["Jueves", "16:30", "Flexiones", 4, "12-15", ""],
    ["Jueves", "16:45", "Fondos en paralelas", 4, "8-10", ""],
    ["Jueves", "17:05", "Dominadas", 5, "6-8", ""],
    ["Lunes", "18:00", "Sentadillas", 4, "20", ""],
    ["Lunes", "18:20", "Plancha (hold)", 3, "30 seg", ""],
]:
    ws.append(r)
style_headers(ws, 6)
autosize(ws, [12, 8, 24, 8, 12, 20])
add_list(ws, DIAS, "A")

# --- Hoja 4: Hitos (pestaña Progreso) ---
ws = wb.create_sheet("Hitos")
ws.append(["Contexto", "Proyecto", "Hito", "Estado", "Orden"])
for r in [
    ["Estudio", "Sprint Diseño", "Investigación inicial", "Hecho", 1],
    ["Estudio", "Sprint Diseño", "Marco teórico", "Hecho", 2],
    ["Estudio", "Sprint Diseño", "Prototipo funcional", "Actual", 3],
    ["Estudio", "Sprint Diseño", "Entrega final", "Pendiente", 4],
    ["Calistenia", "Dominadas", "5 dominadas seguidas", "Hecho", 1],
    ["Calistenia", "Dominadas", "8 dominadas seguidas", "Actual", 2],
    ["Calistenia", "Dominadas", "12 dominadas seguidas", "Pendiente", 3],
]:
    ws.append(r)
style_headers(ws, 5)
autosize(ws, [13, 20, 26, 12, 8])
add_list(ws, CONTEXTOS, "A")
add_list(ws, ESTADOS, "D")

out_dir = os.path.join(os.path.dirname(__file__), "..", "docs")
os.makedirs(out_dir, exist_ok=True)
out_path = os.path.abspath(os.path.join(out_dir, "Plantilla_Rutina.xlsx"))
wb.save(out_path)
print("Plantilla generada en:", out_path)
